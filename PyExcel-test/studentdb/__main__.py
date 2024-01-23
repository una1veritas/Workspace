'''
Created on 2024/01/21

@author: sin
'''
import sys
from openpyxl import Workbook, load_workbook

def main(argv):
    print(argv)
    if len(argv) > 1 :
        fname = argv[1]
        if len(argv) > 2 :
            row_heading = int(argv[2])
        else:
            row_heading = 1
        try:
            wb = load_workbook(filename=fname, data_only=True)
            wsnames = wb.sheetnames
            #print(wb.sheetnames)
        except FileNotFoundError as e:
            print("{} not found.".format(fname))
            exit(1)
        ws = wb[wsnames[0]]
        heading = [ws[cellname].value for cellname in [f"{chr(ord('A')+c)}{row_heading}" for c in range(0,ws.max_column)]]
        column_dict = dict()
        for i in range(len(heading)) :
            column_dict[heading[i]] = i
        #print(heading)
        rows = []
        for row in ws.iter_rows(min_row=row_heading+1,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
            values = []
            for cell in row:
                values.append(cell.value)
            rows.append(values)
            #print(values)
        print(heading)
        rows.sort(key=lambda x: x[column_dict["研究室内発表順"]]) # sort by 研究室内順番
        rows.sort(key=lambda x: x[column_dict["研究室発表順"]]) # sort by 発表順番
        rows.sort(key=lambda x: x[column_dict["発表グループ"]]) # sort by グループ
        for r in rows:
            print(r)
        print(column_dict)
    else:
        wb = Workbook()    
        # grab the active worksheet
        ws = wb.active
        
        # Data can be assigned directly to cells
        ws['A1'] = 42
        
        # Rows can also be appended
        ws.append([1, 2, 3])
        
        # Python types will automatically be converted
        import datetime
        ws['A2'] = datetime.datetime.now()
        
        # Save the file
        wb.save("sample.xlsx")
    
    print("bye.")
    exit(0)
    
if __name__ == '__main__':
    main(sys.argv)
