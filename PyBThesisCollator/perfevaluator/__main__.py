'''
Created on 2024/01/20

@author: Sin Shimozono
'''
import sys, os, glob, re
from openpyxl import Workbook, load_workbook
import xlrd

def read_xlsx(fname, row_heading = 1):
    try:
        wbook = load_workbook(filename=fname, data_only=True)
        wsheetnames = wbook.sheetnames
        #print(wb.sheetnames)
    except FileNotFoundError as e:
        print("{} not found.".format(fname))
        exit(1)
        return None
    wsheet = wbook[wsheetnames[0]]
    heading = list()
    regexpr = re.compile(u'\([0123456789C-]+\)')
    for cellname in [xlrd.colname(c)+"1" for c in range(0, wsheet.max_column)]:
        colname = wsheet[cellname].value
        res = regexpr.search(colname)
        if res is not None :
            colname = res.group()[1:-1]
        heading.append(colname)
    column_dict = dict()
    for i in range(len(heading)) :
        column_dict[heading[i]] = i
    rows = []
    for row in wsheet.iter_rows(min_row=row_heading+1,max_row=wsheet.max_row,min_col=1,max_col=wsheet.max_column):
        values = []
        for cell in row:
            values.append(cell.value)
        rows.append(values)
        #print(values)
    #print(heading)
    return (column_dict, rows)

def db_column(db, col):
    cdict, table = db
    column = list()
    heading = dict()
    if type(col) is str:
        col = cdict[col]
    column.append(table[0][col])
    for row in table[1:] :
        column.append(row[col])
    return column

# def read_csv(filepath):
#     db = list()
#     repatt = re.compile('\([0123456789C-]+\)')
#     with open(filepath, mode="r", encoding='utf-8') as csvf :
#         for a_line in csvf :
#             a_line = a_line.strip()
#             if len(db) == 0:
#                 db.append(list())
#                 for colname in a_line.split(','):
#                     res = repatt.search(colname)
#                     if res is not None :
#                         colname = res.group()[1:-1]
#                     db[-1].append(colname)
#             else:
#                 a_row = a_line.split(',')
#                 db.append(a_row)
#     return db

def db_findrow(db,col,val):
    coldict, table = db
    if col in coldict :
        col = coldict[col]
    res = list()
    for i in range(len(table)):
        if table[i][col] == val :
            res.append(i)
    return res
    
def main(argv):
    filepath = ""
    if len(argv) > 1 :
        filepath = argv[1]
    else:
        print(argv)
        print("No file name.")
        exit(1)
    if filepath.split('.')[-1] != "xlsx" :
        print("File " + filepath + " is not a xlsx.")
        exit(1)        
    if not os.path.isfile(filepath) :
        print("File path " + filepath + " does not exit.")
        exit(1)
    
    db = read_xlsx(filepath)
    print('Reading xlsx data file done.')
    (coldict, table) = db
    heading = [a_pair[0] for a_pair in sorted(coldict.items(), key=lambda x: x[1])]
    sidlist = [col.split('-')[0] for col in heading[5:] if col.endswith('-0')]
    point = {'A':'100', 'B':'90', 'C':'80', 'D':'70', 'E':'60', 'X':'0', None: ''}
    evaltable = list()
    for sid in sidlist:
        r = db_findrow(db, coldict[sid+"-0"], "YES")
        super_row = None
        if len(r) == 0 :
            print("Error: No supervisor for "+sid+".")
            continue # break
        elif len(r) > 1 :
            print("Error: Two or more supervisors for "+sid+".")
            continue # break
        else:
            super_row = r[0]
            # print(sid + " supervised by " + table[super_row][0]+".")        
        evaltable.append([sid])
        evdict = dict()
        for i in range(1,5+1):
            qname = sid+"-"+str(i)
            col = db_column(db, qname)
            val = col.pop(super_row)
            col = [point[idx] for idx in col]
            evdict[sid+"-J"+str(i)] = col
            evdict[sid+"-A"+str(i)] = point[val]
        evaltable[-1].append(evdict)
    print('Re arranging data done.')
    
    filename = os.path.basename(filepath).split(".")
    filename [-1] = 'csv'
    filename = '.'.join(filename)
    print('writing out to ' + filename)
    if not os.path.isdir(os.path.join(".","csvout")) :
        os.mkdir(os.path.join(".","csvout"))
    
    #print(evaltable)
    with open(os.path.join(".","csvout",filename), mode="w", encoding="utf-8") as outf:
        for c in ('sid', 'total', 's-subtotal', 'sQ1', 'sQ2', 'sQ3', 'sQ4', 'sQ5', 'Q1-Q5 subtotal', 'Q1-AVR', 'Q1-J1', 'Q1-J2', ):
            outf.write(c + ',')
        outf.write('\n')
        for sid, evdict in evaltable:
            outf.write(sid + ',,,') # with an empty cells
            keys = sorted(evdict.keys())
            for k in keys[:5]:
                outf.write(evdict[k] + ',')
            outf.write(',,') # supervisor total
            for lk in keys[5:]:
                #print(lk)
                for p in evdict[lk]:
                    outf.write(p + ',')
                outf.write(',') # Qn average
            outf.write('\n')
    print('Bye.')
    
if __name__ == '__main__':
    main(sys.argv)
    