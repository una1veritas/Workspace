'''
Created on 2024/01/20

@author: Sin Shimozono
'''
import sys, os, glob, re
from openpyxl import Workbook, load_workbook

def feedback_opening():
    return u'''<?xml version="1.0" encoding="UTF-8" ?>
<FEEDBACK VERSION="200701" COMMENT="XML-Importfile for mod/feedback">
     <ITEMS>
'''
def feedback_closing():
    return u'''     </ITEMS>
</FEEDBACK>
'''

def feedback_session_header(start_id, gno, tspan, supers):
    tmpstr = u'''          <ITEM TYPE="label" REQUIRED="0">
               <ITEMID><![CDATA[{id_header}]]></ITEMID>
               <ITEMTEXT><![CDATA[]]></ITEMTEXT>
               <ITEMLABEL><![CDATA[]]></ITEMLABEL>
               <PRESENTATION><![CDATA[<h4>グループ{group_no}, {time_span} ({supervisors})</h4>]]></PRESENTATION>
               <OPTIONS><![CDATA[]]></OPTIONS>
               <DEPENDITEM><![CDATA[0]]></DEPENDITEM>
               <DEPENDVALUE><![CDATA[]]></DEPENDVALUE>
          </ITEM>
          <ITEM TYPE="label" REQUIRED="0">
               <ITEMID><![CDATA[{id_evalhint}]]></ITEMID>
               <ITEMTEXT><![CDATA[]]></ITEMTEXT>
               <ITEMLABEL><![CDATA[]]></ITEMLABEL>
               <PRESENTATION><![CDATA[<p><strong><span style="color: #ff0000;">評価値　　A: 100点、B:90点、C:80点、D:70点、E:60点、X:0点</span></strong></p>]]></PRESENTATION>
               <OPTIONS><![CDATA[]]></OPTIONS>
               <DEPENDITEM><![CDATA[0]]></DEPENDITEM>
               <DEPENDVALUE><![CDATA[]]></DEPENDVALUE>
          </ITEM>
'''.format(id_header=start_id, group_no = gno, time_span=tspan,supervisors=supers,id_evalhint= start_id+1)
    return (tmpstr, start_id+2)

def feedback_questions(start_id, sid, sname):
    tmpstr = u'''          <ITEM TYPE="label" REQUIRED="0">
               <ITEMID><![CDATA[{id_qheader}]]></ITEMID>
               <ITEMTEXT><![CDATA[]]></ITEMTEXT>
               <ITEMLABEL><![CDATA[]]></ITEMLABEL>
               <PRESENTATION><![CDATA[<h5>{studentid}&nbsp; &nbsp;{studentname}</h5>]]></PRESENTATION>
               <OPTIONS><![CDATA[]]></OPTIONS>
               <DEPENDITEM><![CDATA[0]]></DEPENDITEM>
               <DEPENDVALUE><![CDATA[]]></DEPENDVALUE>
          </ITEM>
          <ITEM TYPE="multichoice" REQUIRED="0">
               <ITEMID><![CDATA[{id_q0}]]></ITEMID>
               <ITEMTEXT><![CDATA[この学生の指導教員です]]></ITEMTEXT>
               <ITEMLABEL><![CDATA[{studentid}-0]]></ITEMLABEL>
               <PRESENTATION><![CDATA[c>>>>>YES<<<<<1]]></PRESENTATION>
               <OPTIONS><![CDATA[h]]></OPTIONS>
               <DEPENDITEM><![CDATA[0]]></DEPENDITEM>
               <DEPENDVALUE><![CDATA[]]></DEPENDVALUE>
          </ITEM>
          <ITEM TYPE="multichoice" REQUIRED="0">
               <ITEMID><![CDATA[{id_q1}]]></ITEMID>
               <ITEMTEXT><![CDATA[1. 研究成果]]></ITEMTEXT>
               <ITEMLABEL><![CDATA[{studentid}-1]]></ITEMLABEL>
               <PRESENTATION><![CDATA[r>>>>>A|B|C|D|E|X<<<<<1]]></PRESENTATION>
               <OPTIONS><![CDATA[h]]></OPTIONS>
               <DEPENDITEM><![CDATA[0]]></DEPENDITEM>
               <DEPENDVALUE><![CDATA[]]></DEPENDVALUE>
          </ITEM>
          <ITEM TYPE="multichoice" REQUIRED="0">
               <ITEMID><![CDATA[{id_q2}]]></ITEMID>
               <ITEMTEXT><![CDATA[2. 問題解決]]></ITEMTEXT>
               <ITEMLABEL><![CDATA[{studentid}-2]]></ITEMLABEL>
               <PRESENTATION><![CDATA[r>>>>>A|B|C|D|E|X<<<<<1]]></PRESENTATION>
               <OPTIONS><![CDATA[h]]></OPTIONS>
               <DEPENDITEM><![CDATA[0]]></DEPENDITEM>
               <DEPENDVALUE><![CDATA[]]></DEPENDVALUE>
          </ITEM>
          <ITEM TYPE="multichoice" REQUIRED="0">
               <ITEMID><![CDATA[{id_q3}]]></ITEMID>
               <ITEMTEXT><![CDATA[3. コミュニケーション能力]]></ITEMTEXT>
               <ITEMLABEL><![CDATA[{studentid}-3]]></ITEMLABEL>
               <PRESENTATION><![CDATA[r>>>>>A|B|C|D|E|X<<<<<1]]></PRESENTATION>
               <OPTIONS><![CDATA[h]]></OPTIONS>
               <DEPENDITEM><![CDATA[0]]></DEPENDITEM>
               <DEPENDVALUE><![CDATA[]]></DEPENDVALUE>
          </ITEM>
          <ITEM TYPE="multichoice" REQUIRED="0">
               <ITEMID><![CDATA[{id_q4}]]></ITEMID>
               <ITEMTEXT><![CDATA[4. 論理的記述能力]]></ITEMTEXT>
               <ITEMLABEL><![CDATA[{studentid}-4]]></ITEMLABEL>
               <PRESENTATION><![CDATA[r>>>>>A|B|C|D|E|X<<<<<1]]></PRESENTATION>
               <OPTIONS><![CDATA[h]]></OPTIONS>
               <DEPENDITEM><![CDATA[0]]></DEPENDITEM>
               <DEPENDVALUE><![CDATA[]]></DEPENDVALUE>
          </ITEM>
          <ITEM TYPE="multichoice" REQUIRED="0">
               <ITEMID><![CDATA[{id_q5}]]></ITEMID>
               <ITEMTEXT><![CDATA[5. 継続的学習能力（指導教員のみ評価が有効）]]></ITEMTEXT>
               <ITEMLABEL><![CDATA[{studentid}-5]]></ITEMLABEL>
               <PRESENTATION><![CDATA[r>>>>>A|B|C|D|E|X<<<<<1]]></PRESENTATION>
               <OPTIONS><![CDATA[h]]></OPTIONS>
               <DEPENDITEM><![CDATA[{id_q0}]]></DEPENDITEM>
               <DEPENDVALUE><![CDATA[YES]]></DEPENDVALUE>
          </ITEM>
'''.format(id_qheader=start_id, studentid=sid, studentname = sname, 
           id_q0 = start_id + 1, id_q1 = start_id + 2, id_q2 = start_id + 3,
           id_q3 = start_id + 4, id_q4 = start_id + 5, id_q5 = start_id + 7)
    return (tmpstr, start_id+8)


#row_heading 行を heading としてシートの内容を読み出す
def btinfo_db(fname, row_heading = 1):
    try:
        wb = load_workbook(filename=fname, data_only=True)
        wsnames = wb.sheetnames
        #print(wb.sheetnames)
    except FileNotFoundError as e:
        print("{} not found.".format(fname))
        exit(1)
        return None
    ws = wb[wsnames[0]]
    heading = [ws[cellname].value for cellname in [f"{chr(ord('A')+c)}{row_heading}" for c in range(0,ws.max_column)]]
    #print(heading)
    column_dict = dict()
    for i in range(len(heading)) :
        column_dict[heading[i]] = i
    rows = []
    for row in ws.iter_rows(min_row=row_heading+1,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
        values = []
        for cell in row:
            values.append(cell.value)
        rows.append(values)
        #print(values)
    #print(heading)
    return (column_dict, rows)
    
def collect(tbl, col):
    itemset = set()
    result = list()
    for r in tbl:
        item = r[col]
        if not item in itemset:
            itemset.add(item)
            result.append(item)
    return result
    
def write_feedback_xml(group_no, start_time, end_time, supervisor, student_list, start_id = 1000):
    id = start_id
    if not os.path.isdir("xmlout") :
        os.mkdir("xmlout")
    outfilepath = os.path.join("xmlout", "feedback_group{}_{}.xml".format(group_no, supervisor))
    with open(outfilepath, mode="w", encoding="utf-8") as outfile:
        outfile.writelines(feedback_opening())
        (tmpstr, id) = feedback_session_header(id, group_no, start_time+"--"+end_time, supervisor)
        outfile.writelines(tmpstr)
        for st in student_list:
            (resstr, id) = feedback_questions(id, st[3], st[4])
            outfile.writelines(resstr)
        outfile.writelines(feedback_closing())
    return
    
def main(argv):
    btinfofile = "";
    if len(argv) > 1 :
        btinfofile = argv[1]
        if not os.path.isfile(btinfofile) :
            print("File " + btinfofile + " does not exit.")
            exit(1)
    else:
        print(argv)
        print("No info file name.")
        exit(1)

    try:
        (btcolumn, btrows) = btinfo_db(btinfofile, 2)
        #btrows.sort(key=lambda x: x[btcolumn["開始"]]) # sort by 研究室内順番
        #btrows.sort(key=lambda x: x[btcolumn["終了"]]) # sort by 発表順番
        #btrows.sort(key=lambda x: x[btcolumn["グループ"]]) # sort by グループ
        #print(wb.sheetnames)
    except FileNotFoundError as e:
        print(e)
        print("Failed to read workshhet file {}.".format(btinfofile))
        exit(1)
    
    group_list = collect(btrows, btcolumn["グループ"])
    supervisors_list = collect(btrows, btcolumn["指導教員"])
    
    btprogram = list()
    for a_row in btrows:
        group = a_row[btcolumn['グループ']]
        if len(btprogram) == 0:
            btprogram.append( (group, list()) )
        if btprogram[-1][0] != group :
            btprogram.append( (group, list()) )
        s_time = a_row[btcolumn['開始']].strftime("%H:%M")
        e_time = a_row[btcolumn['終了']].strftime("%H:%M")
        s_id = a_row[btcolumn['学生番号']] 
        s_name = a_row[btcolumn['学生氏名']] 
        if a_row[btcolumn['学生番号']] is None :
            continue
        super = a_row[btcolumn['指導教員']]
        if len(btprogram[-1][1]) == 0 :
            btprogram[-1][1].append( (super, list()) )
        if btprogram[-1][1][0] != super :
            btprogram[-1][1].append( (super, list()) )
        btprogram[-1][1].append([s_id, s_name])

    print(btprogram)
    exit(0)
    for a_line in feedback_questions(1000,"123C5678","川津花子")[0].split("\n"):
        print(a_line)
#    with open("tempout.csv", mode="w", encoding="utf-8") as outcsvf:
        
    print("bye.")
    
    exit(0)


if __name__ == '__main__':
    main(sys.argv)