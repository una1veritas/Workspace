'''
Created on 2024/01/20

@author: Sin Shimozono
'''
import sys, os, glob, re
import pypdf
from openpyxl import Workbook, load_workbook
from operator import itemgetter
import shutil

def pdffile_list(dirpath) :
    dirpath = os.path.join(dirpath, u"**", u"*.pdf")
    paths = glob.glob(dirpath, recursive=True)
    if len(paths) == 0 :
        print("no pdf files found.")
        return None

    db = list()
    for filepath in paths :
        filename = os.path.basename(filepath)
        if not filename.endswith(".pdf") :
            # not a pdf file.
            continue
        parts = os.path.splitext(filename)[0].split('_')
        if len(parts) > 4 :
            print("warning: too many _'s in file name: " + str(parts))
        elif len(parts) < 3 :
            print("error: too few _'s in file name: " + str(parts))
            exit(1)
            return None
        (sid, sname, supervisor) = parts[0:3]
        db.append([sid, sname, supervisor, filepath])
        #print(db[-1])
    return db
    #

#row_heading 行を heading としてシートの内容を読み出す
def btinfo_db(fname, row_heading = 1):
    try:
        wb = load_workbook(filename=fname, data_only=True)
        wsnames = wb.sheetnames
        #print(wb.sheetnames)
    except FileNotFoundError as e:
        print("{} not found.".format(fname))
        exit(1)
        return None
    ws = wb[wsnames[0]]
    heading = [ws[cellname].value for cellname in [f"{chr(ord('A')+c)}{row_heading}" for c in range(0,ws.max_column)]]
    #print(heading)
    column_dict = dict()
    for i in range(len(heading)) :
        column_dict[heading[i]] = i
    rows = []
    for row in ws.iter_rows(min_row=row_heading+1,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
        values = []
        for cell in row:
            values.append(cell.value)
        rows.append(values)
        #print(values)
    #print(heading)
    return (column_dict, rows)
    
def main(argv):
    pdfbasedir = u"."
    btinfofile = "";
    if argv[1].endswith(".xlsx") :
        btinfofile = argv[1]
        pdfbasedir = argv[2]
    elif argv[2].endswith(".xlsx") :
        btinfofile = argv[2]
        pdfbasedir = argv[1]
    else:
        print(argv)
        print("No info file name.")
        exit(1)
    if not os.path.isdir(pdfbasedir) :
        print("Directory " + pdfbasedir + " does not exit.")
        exit(1)
    try:
        (btcolumn, btrows) = btinfo_db(btinfofile, 2)
        btrows.sort(key=lambda x: x[btcolumn["研究室内発表順"]]) # sort by 研究室内順番
        btrows.sort(key=lambda x: x[btcolumn["研究室発表順"]]) # sort by 発表順番
        btrows.sort(key=lambda x: x[btcolumn["発表グループ"]]) # sort by グループ
        #print(wb.sheetnames)
        # print(btcolumn)
        # for r in btrows:
        #     print(r)
        # print()
    except FileNotFoundError as e:
        print(e)
        print("Failed to read workshhet file {}.".format(btinfofile))
        exit(1)
    
    renamedbtfilesdir = "bts"
    writer = pypdf.PdfWriter()
    pdfdb = pdffile_list(pdfbasedir)
    os.mkdir(renamedbtfilesdir)
    for entry in btrows :
        if len(entry[8]) == 0 :
            # no title
            continue
        progno = "{}-{}-{}-".format(entry[-3], entry[-2], entry[-1])
        pdffilename = next((each[3] for each in pdfdb if each[0] == entry[0]), None)
        if pdffilename == None:
            print("error!! No correspoinding file name.")
            continue
        
        reader = pypdf.PdfReader(pdffilename)
        # テキスト抽出
        nofpages = len(reader.pages)
        if nofpages > 1 :
            shutil.copy2(pdffilename, os.path.join(renamedbtfilesdir,progno+os.path.basename(pdffilename)) )
            # titlepagetext = reader.pages[0].extract_text()
            # print(titlepagetext)
            pdfpage = reader.pages[1]
            writer.add_page(pdfpage)
            print(pdffilename + ", " + str(nofpages) + "pages.", end="")
        else:
            print()
            print("error!! The document has insufficient pages.")
        print() 
    writer.write("ext-abstracts.pdf")
    print(str(len(pdfdb)) + " files in " + pdfbasedir)
    print("bye.")
    
    exit(0)


if __name__ == '__main__':
    main(sys.argv)