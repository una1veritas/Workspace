'''
Created on 2024/01/21

@author: sin
'''
import sys
from openpyxl import Workbook, load_workbook

def main(argv):
    print(argv)
    if len(argv) >= 1 :
        fname = argv[1]
        try:
            wb = load_workbook(filename=fname, data_only=True)
            wsnames = wb.sheetnames
            #print(wb.sheetnames)
        except FileNotFoundError as e:
            print("{} not found.".format(fname))
            exit(1)
        ws = wb[wsnames[0]]
        heading = [ws[cellname].value for cellname in [f"{chr(ord('A')+c)}3" for c in range(0,ws.max_column)]]
        #print(heading)
        rows = []
        for row in ws.iter_rows(min_row=4,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
            values = []
            for cell in row:
                values.append(cell.value)
            rows.append(values)
            #print(values)
        print(heading)
        rows.sort(key=lambda x: x[1]) # sort by 発表順番
        rows.sort(key=lambda x: x[0]) # sort by グループ
        for row in rows:
            print(row)
    else:
        wb = Workbook()    
        # grab the active worksheet
        ws = wb.active
        
        # Data can be assigned directly to cells
        ws['A1'] = 42
        
        # Rows can also be appended
        ws.append([1, 2, 3])
        
        # Python types will automatically be converted
        import datetime
        ws['A2'] = datetime.datetime.now()
        
        # Save the file
        wb.save("sample.xlsx")
    
    print("bye.")
    exit(0)
    
if __name__ == '__main__':
    main(sys.argv)
